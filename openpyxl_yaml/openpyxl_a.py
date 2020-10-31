#!/usr/bin/env python
# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook
import random
import time
import os

'''
# 创建表单
wb = Workbook()  # 实例化
ws1 = wb.active  # 默认创建获取第一个工作表的索引对象
ws2 = wb.create_sheet('mysheet')  # 默认插入到最后，新建一个sheet对象
ws3 = wb.create_sheet('onesheet',0)  # 通过索引改变插入位置
wb.create_sheet()   # 新建一个sheet 名称不指定的话默认Sheet Sheet1 ...创建
ws4 = wb['Sheet1']  # 通过名称所谓索引指定sheet的引用
ws4.title = 'tests'  # 重命名
print(wb.sheetnames)
------------------------------------------------------------------------
# 读取文件                  
wb = load_workbook(path)
ws = wb[sheetname]      
------------------------------------------------------------------------
# 访问单元格
d = ws.cell(row=4, column=2, value=10)
# 创建访问单元格并赋值
for x in range(1,101):  
     for y in range(1,101):
 		ws.cell(row=x,colum=y)
# 批量访问单元格对象
# 所有行 ws.rows
# 所有列 ws.colums
# 赋值 c.value = value
-------------------------------------------------------------------------
# 保存文件
# 普通文件 wb.save(name)
# 流文件 wb.save(name,as_template=True)
-------------------------------------------------------------------------
'''


def _phone():
	return f'199{random.randint(10000000, 999999999)}'


def _time():
	return time.strftime('%Y/%m/%d', time.localtime(int(round(time.time())) * random.random()))


def _email():
	return f'{random.randint(10000, 99999)}@bestcem.com'


class Excel:
	'''
	排序str："text{}".format(j)
	固定str："str"
	int: random.randomint(start,end)
	date: _time
	phone: _phone
	email: _email
	float：random.uniform(start,end)
	'''

	def read_excel(self, path, sheetname='Sheet1'):
		wb = load_workbook(path)
		ws = wb[sheetname]
		l = []
		for row in ws.values:
			l.append(row)
		return l

	def write_excel(self, read_path, write_path, num, sheetname='Sheet'):
		wb = Workbook()
		ws = wb.active
		ws.name = sheetname
		demo = self.read_excel(read_path)
		title = demo[0]
		data_type = demo[1]
		for i in range(1, len(title) + 1):  # 将title写入sheet
			ws.cell(row=1, column=i, value=title[i - 1])
		for i in range(len(data_type)):  # 写测试数据
			if data_type[i]:
				for j in range(num):
					ws.cell(row=j + 2, column=i + 1, value=eval(data_type[i]))
		wb.save(write_path)


if __name__ == '__main__':
	Excel().write_excel('demo.xlsx', 'demo1.xlsx',10)

